from PyPDF2 import PdfWriter, PdfReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook, load_workbook

# set font
pdfmetrics.registerFont(TTFont('F1', 'THSarabunNew.ttf'))

from datetime import date




    
def fillData(row):
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=A4)

    '''
   {'prefix': 'เด็กหญญิง', 'name': 'สมใจ', 'surname': 'ระยับศรี', 'dob': '11/04/2554', 'id_number': '1103500102376', 'home_number': '0', 'moo': '0', 'tumbon': 'ชากพง', 'amphur': 'แกลง', 'provice': 'ระยอง', 'class': 'ป.6'}
    '''

    # name
    can.setFont('F1', 16)
    can.drawString(150, 602.5, row.get('prefix') + row.get('name') + " " + row.get('surname'))

    # date
    can.drawString(370, 602.5, row.get('dob').split('/')[0])

    # month
 
    can.drawString(460, 602.5, getMonth(int(row.get('dob').split('/')[1])))

    # year budhist
    can.drawString(93, 581, row.get('dob').split('/')[2])

    # age
    can.drawString(158, 581, str((date.today().year + 543) - int(row.get('dob').split('/')[2])))

    # id_number
    id = row.get('id_number')
    start_position = 314.5
    for i in id:
        can.drawString(start_position, 579, str(i))
        start_position += 18

    # address
    
    can.drawString(134, 560.5, row.get('home_number') if row.get('home_number') != '0' else "")
    can.drawString(193, 560.5, row.get('moo') if row.get('moo') != '0' else "")
    can.drawString(256, 560.5, row.get('tumbon') if row.get('tumbon') != '0' else "")
    can.drawString(365, 560.5, row.get('amphur') if row.get('amphur') != '0' else "")
    can.drawString(468, 560.5, row.get('province') if row.get('province') != '0' else "")

    # class
    can.drawString(157, 539.5, row.get('class'))

    # school
    can.drawString(248, 539.5, "วัดคลองชากพง")

    # sport type
    can.drawString(242, 497.5, "วอลเลย์บอล")

    # age ver
    can.drawString(417, 497.5, "12")

    # sign
    can.drawString(278, 430, row.get('prefix') + row.get('name') + " " + row.get('surname'))

    ######################## คำรับรองโรงเรียน ###########################
    # school
    can.drawString(177, 329, "โรงเรียนวัดคลองชากพง")
    # Amphur
    can.drawString(370, 329, "แกลง")
    # dob
    can.drawString(435, 308, row.get('dob').split('/')[0])
    can.drawString(487, 308, getMonth(int(row.get('dob').split('/')[1])))
    can.drawString(98, 286.8, row.get('dob').split('/')[2])
    # class
    can.drawString(250, 286.8, row.get('class'))

    # sign head quater
    can.drawString(290, 183.5, "นางอรอนงค์ ผลทวีวัฒนชัย")
    can.drawString(265, 162.5, "รักษาการในตำแหน่งผู้อำนวยการโรงเรียนวัดคลองชากพง")

    can.save()

    #move to the beginning of the StringIO buffer
    packet.seek(0)

    # create a new PDF with Reportlab
    new_pdf = PdfReader(packet)
    # read your existing PDF
    existing_pdf = PdfReader(open("subscription.pdf", "rb"))
    output = PdfWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.pages[0]
    page.merge_page(new_pdf.pages[0])
    output.add_page(page)
    # finally, write "output" to a real file
    output_stream = open("output/{}.pdf".format(row.get('name')), "wb")
    output.write(output_stream)
    output_stream.close()

def readXL():
    wb = load_workbook('data.xlsx')
    ws = wb['Sheet1']
    cell_range = ws['A1:K20']

    data = [{k.value:v.value for k, v in zip([r for r in cell_range][0], r)} for r in cell_range]

    del data[0]
    return data

def getMonth(num):
    if num == 1:
        return "มกราคม"
    elif num == 2:
        return "กุมภาพันธ์"
    elif num == 3:
        return "มีนาคม"
    elif num == 4:
        return "เมษายน "
    elif num == 5:
        return "พฤษภาคม"
    elif num == 6:
        return "มิถุนายน"
    elif num == 7:
        return "กรกฎาคม"
    elif num == 8:
        return "สิงหาคม"
    elif num == 9:
        return "กันยายน"
    elif num == 10:
        return "ตุลาคม"
    elif num == 11:
        return "พฤศจิกายน"
    elif num == 12:
        return "ธันวาคม"

if __name__ == '__main__':
    data = readXL()
    
    for d in data:
        print(d)
        fillData(d)